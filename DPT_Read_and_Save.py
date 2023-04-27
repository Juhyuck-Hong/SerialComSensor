import serial
import struct
import crcmod
from time import sleep
import time
import datetime
from openpyxl import Workbook

# create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# write the headers to the worksheet
ws.cell(row=1, column=1, value="Time")
ws.cell(row=1, column=2, value="Measure Value")
ws.cell(row=1, column=3, value="Measure Value2")

# list of data to save
data_p = []
data_m1 = []
data_m2 = []

# make serial connection
ser1 = serial.Serial('COM5', 9600, parity='N', stopbits=1, bytesize=8)
ser2 = serial.Serial('COM6', 9600, parity='N', stopbits=1, bytesize=8)

# iteration count
repeat = 0

# maximum repetition
maxRepeat = 50

# average span (sec)
averageSpan = 10

# except when KeyboardInterrupt
try:
    # set time checking start point
    start_time = time.time()
    while True:
        # time stamp
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # modbus protocol of pressure sensor
        command_dict = {'02': "pressure_unit",
                        '03': "decimal_point",
                        '04': "measured output",
                        '05': "transmitter zero point",
                        '06': "transmitter extreme point"}
        decimal_point = {0: 1,
                         1: 0.1,
                         2: 0.01,
                         3: 0.001}
        pressure_unit = {0: "MPa",
                         1: "kPa",
                         2: "Pa",
                         3: "bar",
                         4: "mbar",
                         5: "kg/㎠",
                         6: "psi",
                         7: "mH2O",
                         8: "mmH2O"}

        # repeat for each command
        for i in command_dict.keys():
            # Calculate the CRC16 value for the Modbus RTU message
            crc16 = crcmod.predefined.Crc('modbus')

            # initiate Tx codes
            machine_addr = b'\x01'
            function_code = b'\x03'
            data_start_addr = b'\x00'
            data_Nos = b'\x00\x01'
            data_send = machine_addr + function_code + \
                data_start_addr + bytes.fromhex(i) + data_Nos

            # print(data_send)
            crc16.update(data_send)
            crc_bytes = crc16.digest()

            # Reverse the byte order of the CRC16 bytes
            crc_bytes_reversed = crc_bytes[::-1]

            # Append the CRC16 value to the Modbus RTU message
            data = data_send + crc_bytes_reversed

            # Send the Modbus RTU request
            ser1.write(data)
            ser2.write(data)

            # Wait for a response
            response1 = ser1.read(7)
            response2 = ser2.read(7)

            # check Tx and Rx Hex
            # print(f"[-100~100Pa] Tx: {data}, Rx: {response1}")
            # print(f"[-1 ~ 1 kPa] Tx: {data}, Rx: {response2}")

            # Extract the relevant fields
            _, _, _, value1, crc_code1 = struct.unpack('>BBBhH', response1)
            _, _, _, value2, crc_code2 = struct.unpack('>BBBhH', response2)

            # Print the extracted fields
            if command_dict[i] == "decimal_point":
                dp1 = decimal_point[value1]
                dp2 = decimal_point[value2]
                # print(f"[-100~100Pa] {command_dict[i]}: {value1}, ({dp1})")
                # print(f"[-1 ~ 1 kPa] {command_dict[i]}: {value2}, ({dp2})")

            elif command_dict[i] == "pressure_unit":
                pu1 = pressure_unit[value1]
                pu2 = pressure_unit[value2]
                # print(f"[-100~100Pa] {command_dict[i]}: {value1}, ({pu1})")
                # print(f"[-1 ~ 1 kPa] {command_dict[i]}: {value2}, ({pu2})")

            elif command_dict[i] == "measured output":
                m_value1 = value1
                m_value2 = value2
                # print(f"[-100~100Pa] {command_dict[i]}: {value1}")
                # print(f"[-1 ~ 1 kPa] {command_dict[i]}: {value2}")

            else:
                # print(f"[-100~100Pa] {command_dict[i]}: {value1}")
                # print(f"[-1 ~ 1 kPa] {command_dict[i]}: {value2}")
                pass

        # print(f"measured 1: {m_value1 * dp1} {pu1} ({m_value1}, {dp1})")
        # print(f"measured 2: {m_value2 * dp2} {pu2} ({m_value2}, {dp2})")

        # save measured data with calculation
        m1 = f"{m_value1 * dp1: ,.0f}"
        m2 = f"{m_value2 * dp2 * 1000: ,.0f}"
        mdiff = f"{m_value1*dp1 - m_value2 * dp2 * 1000: ,.0f}"
        data_m1.append(float(m1))
        data_m2.append(float(m2))

        elapsed_time = time.time() - start_time

        # make average for some secs
        if elapsed_time >= averageSpan:
            avg_m1 = sum(data_m1) / len(data_m1)
            avg_m2 = sum(data_m2) / len(data_m2)
            mdiff = avg_m1 - avg_m2
            print(
                f"[{repeat}] M1: {avg_m1: ,.2f} Pa | M2: {avg_m2: ,.2f} Pa (Δ: {mdiff: ,.2f})")
            data_p.append(f"{timestamp},{avg_m1},{avg_m2}")
            data_m1 = []
            data_m2 = []
            start_time = time.time()
            repeat += 1

        # raise KeyboardInterrupt when the iteration count is completed
        if repeat > maxRepeat:
            raise KeyboardInterrupt

# when terminating
except KeyboardInterrupt:
    # serial connection closing
    ser1.close()
    ser2.close()
    # separates data
    pressure_list = [i.split(',')[1] for i in data_p]
    pressure_list += [i.split(',')[2] for i in data_p]
    time_list = [i.split(',')[0] for i in data_p]

    # time diff check in min.
    time1 = datetime.datetime.strptime(min(time_list), '%Y-%m-%d %H:%M:%S')
    time2 = datetime.datetime.strptime(max(time_list), '%Y-%m-%d %H:%M:%S')
    timestamp1 = int(time1.timestamp())
    timestamp2 = int(time2.timestamp())
    diff_seconds = abs(timestamp2 - timestamp1)
    diff_minutes = diff_seconds // 60

    # write the data to the worksheet
    for i, row in enumerate(data_p, start=2):
        row = row.split(',')
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=float(row[1]))
        ws.cell(row=i, column=3, value=float(row[2]))

    # save the workbook to an Excel file (file name = min Pa ~ max Pa _ time diff in min)
    # timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    p_min = float(min(pressure_list))
    p_max = float(max(pressure_list))
    wb.save(f"({p_min: ,.2f} ~ {p_max: ,.2f})_{diff_minutes}min.xlsx")

    print("stopped")
